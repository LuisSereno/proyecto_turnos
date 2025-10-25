"""
EXPORTACION.PY - Utilidades para exportación de planillas
"""

from io import BytesIO
from datetime import datetime, timedelta
import json
import csv

# Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# PDF
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas

    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# iCalendar
try:
    from icalendar import Calendar, Event

    ICAL_AVAILABLE = True
except ImportError:
    ICAL_AVAILABLE = False

import logging

logger = logging.getLogger(__name__)


def generar_excel_planilla(ejecucion):
    """
    Genera archivo Excel con la planilla de turnos

    Args:
        ejecucion: Instancia de EjecucionPlanificacion

    Returns:
        BytesIO: Buffer con el contenido del Excel
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl no está instalado. Ejecuta: pip install openpyxl")

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Planilla de Turnos"

        # Estilos
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)

        turno_manana_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
        turno_tarde_fill = PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid")
        turno_noche_fill = PatternFill(start_color="6F42C1", end_color="6F42C1", fill_type="solid")

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título
        ws['A1'] = f"Planificación: {ejecucion.configuracion.nombre}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        ws['A2'] = f"Período: {ejecucion.configuracion.num_dias} días desde {ejecucion.configuracion.fecha_inicio}"
        ws.merge_cells('A2:D2')

        ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws.merge_cells('A3:D3')

        # Espacio
        current_row = 5

        # Headers
        ws[f'A{current_row}'] = 'Día'
        ws[f'B{current_row}'] = 'Fecha'
        ws[f'C{current_row}'] = 'Turno'
        ws[f'D{current_row}'] = 'Enfermeras'

        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{current_row}']
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        current_row += 1

        # Datos
        planilla = ejecucion.planilla or {}
        fecha_inicio = ejecucion.configuracion.fecha_inicio

        for i in range(1, ejecucion.configuracion.num_dias + 1):
            dia_key = f"dia_{i}"
            fecha_actual = fecha_inicio + timedelta(days=i - 1)
            turnos = planilla.get(dia_key, {})

            for turno_tipo in ['MAÑANA', 'TARDE', 'NOCHE']:
                enfermeras = turnos.get(turno_tipo, [])

                ws[f'A{current_row}'] = i
                ws[f'B{current_row}'] = fecha_actual.strftime('%d/%m/%Y')
                ws[f'C{current_row}'] = turno_tipo
                ws[f'D{current_row}'] = ', '.join(enfermeras) if enfermeras else 'Sin asignar'

                # Aplicar color según turno
                if turno_tipo == 'MAÑANA':
                    ws[f'C{current_row}'].fill = turno_manana_fill
                elif turno_tipo == 'TARDE':
                    ws[f'C{current_row}'].fill = turno_tarde_fill
                elif turno_tipo == 'NOCHE':
                    ws[f'C{current_row}'].fill = turno_noche_fill

                # Bordes
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{current_row}'].border = border

                current_row += 1

        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 50

        # Añadir hoja de estadísticas
        ws_stats = wb.create_sheet("Estadísticas")
        ws_stats['A1'] = "Estadísticas de la Planificación"
        ws_stats['A1'].font = Font(size=14, bold=True)

        ws_stats['A3'] = "Penalización Total:"
        ws_stats['B3'] = ejecucion.penalizacion_total or 0

        ws_stats['A4'] = "Es Óptima:"
        ws_stats['B4'] = "Sí" if ejecucion.es_optima else "No"

        ws_stats['A5'] = "Duración (segundos):"
        ws_stats['B5'] = ejecucion.duracion or 0

        ws_stats['A6'] = "Estado:"
        ws_stats['B6'] = ejecucion.estado

        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        logger.info(f"Excel generado exitosamente para ejecución {ejecucion.id}")
        return buffer

    except Exception as e:
        logger.error(f"Error al generar Excel: {str(e)}")
        raise


def generar_pdf_planilla(ejecucion):
    """
    Genera archivo PDF con la planilla de turnos

    Args:
        ejecucion: Instancia de EjecucionPlanificacion

    Returns:
        BytesIO: Buffer con el contenido del PDF
    """
    if not PDF_AVAILABLE:
        raise ImportError("reportlab no está instalado. Ejecuta: pip install reportlab")

    try:
        buffer = BytesIO()

        # Crear documento
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=18
        )

        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#667eea'),
            spaceAfter=30,
            alignment=1  # Center
        )

        # Contenido
        story = []

        # Título
        title = Paragraph(f"Planificación: {ejecucion.configuracion.nombre}", title_style)
        story.append(title)

        # Información
        info_style = styles['Normal']
        info_text = f"""
        <b>Período:</b> {ejecucion.configuracion.num_dias} días desde {ejecucion.configuracion.fecha_inicio.strftime('%d/%m/%Y')}<br/>
        <b>Generado:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}<br/>
        <b>Estado:</b> {ejecucion.estado}<br/>
        <b>Penalización:</b> {ejecucion.penalizacion_total or 0}
        """
        story.append(Paragraph(info_text, info_style))
        story.append(Spacer(1, 20))

        # Tabla de datos
        planilla = ejecucion.planilla or {}
        fecha_inicio = ejecucion.configuracion.fecha_inicio

        # Headers
        data = [['Día', 'Fecha', 'Turno', 'Enfermeras']]

        # Datos
        for i in range(1, ejecucion.configuracion.num_dias + 1):
            dia_key = f"dia_{i}"
            fecha_actual = fecha_inicio + timedelta(days=i - 1)
            turnos = planilla.get(dia_key, {})

            for turno_tipo in ['MAÑANA', 'TARDE', 'NOCHE']:
                enfermeras = turnos.get(turno_tipo, [])
                data.append([
                    str(i),
                    fecha_actual.strftime('%d/%m/%Y'),
                    turno_tipo,
                    ', '.join(enfermeras) if enfermeras else 'Sin asignar'
                ])

        # Crear tabla
        table = Table(data, colWidths=[50, 100, 100, 400])

        # Estilo de tabla
        table_style = TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Datos
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ])

        table.setStyle(table_style)
        story.append(table)

        # Construir PDF
        doc.build(story)
        buffer.seek(0)

        logger.info(f"PDF generado exitosamente para ejecución {ejecucion.id}")
        return buffer

    except Exception as e:
        logger.error(f"Error al generar PDF: {str(e)}")
        raise


def generar_csv_planilla(ejecucion):
    """
    Genera archivo CSV con la planilla de turnos

    Args:
        ejecucion: Instancia de EjecucionPlanificacion

    Returns:
        BytesIO: Buffer con el contenido del CSV
    """
    try:
        buffer = BytesIO()

        # Usar TextIOWrapper para manejar encoding
        import io
        text_buffer = io.TextIOWrapper(buffer, encoding='utf-8-sig', newline='')

        writer = csv.writer(text_buffer, delimiter=';', quoting=csv.QUOTE_MINIMAL)

        # Headers
        writer.writerow(['Día', 'Fecha', 'Turno', 'Enfermeras'])

        # Datos
        planilla = ejecucion.planilla or {}
        fecha_inicio = ejecucion.configuracion.fecha_inicio

        for i in range(1, ejecucion.configuracion.num_dias + 1):
            dia_key = f"dia_{i}"
            fecha_actual = fecha_inicio + timedelta(days=i - 1)
            turnos = planilla.get(dia_key, {})

            for turno_tipo in ['MAÑANA', 'TARDE', 'NOCHE']:
                enfermeras = turnos.get(turno_tipo, [])
                writer.writerow([
                    i,
                    fecha_actual.strftime('%d/%m/%Y'),
                    turno_tipo,
                    ', '.join(enfermeras) if enfermeras else 'Sin asignar'
                ])

        text_buffer.flush()
        buffer.seek(0)

        logger.info(f"CSV generado exitosamente para ejecución {ejecucion.id}")
        return buffer

    except Exception as e:
        logger.error(f"Error al generar CSV: {str(e)}")
        raise


def generar_json_planilla(ejecucion):
    """
    Genera archivo JSON con la planilla de turnos

    Args:
        ejecucion: Instancia de EjecucionPlanificacion

    Returns:
        BytesIO: Buffer con el contenido del JSON
    """
    try:
        data = {
            'configuracion': {
                'id': ejecucion.configuracion.id,
                'nombre': ejecucion.configuracion.nombre,
                'num_dias': ejecucion.configuracion.num_dias,
                'fecha_inicio': ejecucion.configuracion.fecha_inicio.isoformat(),
            },
            'ejecucion': {
                'id': ejecucion.id,
                'estado': ejecucion.estado,
                'fecha_inicio': ejecucion.fecha_inicio.isoformat() if ejecucion.fecha_inicio else None,
                'fecha_fin': ejecucion.fecha_fin.isoformat() if ejecucion.fecha_fin else None,
                'duracion': ejecucion.duracion,
                'penalizacion_total': ejecucion.penalizacion_total,
                'es_optima': ejecucion.es_optima,
            },
            'planilla': ejecucion.planilla or {},
            'mensajes': ejecucion.mensajes or {},
            'generado': datetime.now().isoformat(),
        }

        json_str = json.dumps(data, indent=2, ensure_ascii=False)
        buffer = BytesIO(json_str.encode('utf-8'))
        buffer.seek(0)

        logger.info(f"JSON generado exitosamente para ejecución {ejecucion.id}")
        return buffer

    except Exception as e:
        logger.error(f"Error al generar JSON: {str(e)}")
        raise


def generar_ical_planilla(ejecucion):
    """
    Genera archivo iCalendar (.ics) con los turnos

    Args:
        ejecucion: Instancia de EjecucionPlanificacion

    Returns:
        BytesIO: Buffer con el contenido del iCal
    """
    if not ICAL_AVAILABLE:
        raise ImportError("icalendar no está instalado. Ejecuta: pip install icalendar")

    try:
        cal = Calendar()
        cal.add('prodid', '-//Planificador de Turnos//ES')
        cal.add('version', '2.0')
        cal.add('calscale', 'GREGORIAN')
        cal.add('method', 'PUBLISH')
        cal.add('x-wr-calname', f'Turnos - {ejecucion.configuracion.nombre}')
        cal.add('x-wr-timezone', 'Europe/Madrid')
        cal.add('x-wr-caldesc', f'Planificación de turnos generada el {datetime.now().strftime("%d/%m/%Y")}')

        planilla = ejecucion.planilla or {}
        fecha_inicio = ejecucion.configuracion.fecha_inicio

        # Horarios de turnos
        horarios = {
            'MAÑANA': ('07:00', '15:00'),
            'TARDE': ('15:00', '23:00'),
            'NOCHE': ('23:00', '07:00')
        }

        for i in range(1, ejecucion.configuracion.num_dias + 1):
            dia_key = f"dia_{i}"
            fecha_actual = fecha_inicio + timedelta(days=i - 1)
            turnos = planilla.get(dia_key, {})

            for turno_tipo, (hora_inicio, hora_fin) in horarios.items():
                enfermeras = turnos.get(turno_tipo, [])

                if enfermeras:
                    event = Event()
                    event.add('summary', f'Turno {turno_tipo}')

                    # Calcular datetime de inicio y fin
                    inicio_hora, inicio_min = map(int, hora_inicio.split(':'))
                    fin_hora, fin_min = map(int, hora_fin.split(':'))

                    dt_inicio = datetime.combine(fecha_actual,
                                                 datetime.min.time().replace(hour=inicio_hora, minute=inicio_min))

                    # Si el turno termina al día siguiente (NOCHE)
                    if fin_hora < inicio_hora:
                        dt_fin = datetime.combine(fecha_actual + timedelta(days=1),
                                                  datetime.min.time().replace(hour=fin_hora, minute=fin_min))
                    else:
                        dt_fin = datetime.combine(fecha_actual,
                                                  datetime.min.time().replace(hour=fin_hora, minute=fin_min))

                    event.add('dtstart', dt_inicio)
                    event.add('dtend', dt_fin)
                    event.add('description', f'Enfermeras asignadas:\n' + '\n'.join(enfermeras))
                    event.add('location', 'Hospital')
                    event.add('status', 'CONFIRMED')

                    cal.add_component(event)

        buffer = BytesIO(cal.to_ical())
        buffer.seek(0)

        logger.info(f"iCal generado exitosamente para ejecución {ejecucion.id}")
        return buffer

    except Exception as e:
        logger.error(f"Error al generar iCal: {str(e)}")
        raise


def exportar_enfermeras_excel(enfermeras_queryset):
    """
    Exporta lista de enfermeras a Excel

    Args:
        enfermeras_queryset: QuerySet de Enfermera

    Returns:
        BytesIO: Buffer con el contenido del Excel
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl no está instalado")

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Enfermeras"

        # Headers
        headers = ['ID', 'Nombre', 'Email', 'Teléfono', 'DNI', 'Fecha Alta', 'Activa']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        # Datos
        for row_num, enfermera in enumerate(enfermeras_queryset, 2):
            ws.cell(row=row_num, column=1, value=enfermera.id)
            ws.cell(row=row_num, column=2, value=enfermera.nombre)
            ws.cell(row=row_num, column=3, value=enfermera.email)
            ws.cell(row=row_num, column=4, value=enfermera.telefono or '')
            ws.cell(row=row_num, column=5, value=enfermera.dni or '')
            ws.cell(row=row_num, column=6,
                    value=enfermera.fecha_alta.strftime('%d/%m/%Y') if enfermera.fecha_alta else '')
            ws.cell(row=row_num, column=7, value='Sí' if enfermera.activa else 'No')

        # Ajustar anchos
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer

    except Exception as e:
        logger.error(f"Error al exportar enfermeras a Excel: {str(e)}")
        raise
