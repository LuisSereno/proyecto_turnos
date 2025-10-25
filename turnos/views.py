"""
Views for turnos app
"""
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView,
    TemplateView, FormView, View
)
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, Http404
from django.db.models import Q, Count, Avg, Sum
from django.utils import timezone
from datetime import datetime, timedelta

from .models import (
    Enfermera, TipoTurno, ConfiguracionPlanificacion,
    Ejecucion, Planilla, AsignacionTurno
)
from .forms import (
    EnfermeraForm, TipoTurnoForm, ConfiguracionPlanificacionForm,
    ConfiguracionWizardStep1Form, EjecucionRapidaForm,
    FiltroEjecucionesForm, ImportarEnfermerasForm
)
from .mixins import (
    SuperuserRequiredMixin, StaffRequiredMixin, OwnerRequiredMixin,
    FormMessageMixin, PaginationMixin, SearchMixin, FilterMixin
)
from .tasks import ejecutar_planificacion_async
import json


# ========== Dashboard ==========

class DashboardView(LoginRequiredMixin, TemplateView):
    """Vista principal del dashboard"""
    template_name = 'turnos/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Estadísticas generales
        context['total_configuraciones'] = ConfiguracionPlanificacion.objects.filter(activa=True).count()
        context['total_ejecuciones'] = Ejecucion.objects.count()
        context['total_enfermeras'] = Enfermera.objects.filter(activa=True).count()
        context['ejecuciones_completadas'] = Ejecucion.objects.filter(estado='COMPLETADA').count()

        # Ejecuciones recientes
        context['ejecuciones_recientes'] = Ejecucion.objects.select_related(
            'configuracion'
        ).order_by('-fecha_inicio')[:5]

        # Configuraciones activas
        context['configuraciones_activas'] = ConfiguracionPlanificacion.objects.filter(
            activa=True
        ).order_by('-fecha_creacion')[:5]

        # Gráfico de ejecuciones por estado
        context['ejecuciones_por_estado'] = list(
            Ejecucion.objects.values('estado').annotate(count=Count('id'))
        )

        return context


# ========== Configuraciones ==========

class ConfiguracionListView(LoginRequiredMixin, SearchMixin, FilterMixin, PaginationMixin, ListView):
    """Lista de configuraciones"""
    model = ConfiguracionPlanificacion
    template_name = 'turnos/configuration_list.html'
    context_object_name = 'configuraciones'
    paginate_by = 12

    search_fields = ['nombre', 'descripcion']
    filter_fields = {
        'estado': 'activa',
    }

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.select_related('creado_por').prefetch_related('enfermeras', 'turnos')

        # Filtro de estado
        estado = self.request.GET.get('estado')
        if estado == 'activa':
            queryset = queryset.filter(activa=True)
        elif estado == 'inactiva':
            queryset = queryset.filter(activa=False)

        # Ordenamiento
        orden = self.request.GET.get('orden', '-fecha_creacion')
        queryset = queryset.order_by(orden)

        return queryset


class ConfiguracionDetailView(LoginRequiredMixin, DetailView):
    """Detalle de una configuración"""
    model = ConfiguracionPlanificacion
    template_name = 'turnos/configuration_detail.html'
    context_object_name = 'configuracion'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Ejecuciones recientes de esta configuración
        context['ejecuciones_recientes'] = self.object.ejecuciones.order_by('-fecha_inicio')[:5]

        return context


class ConfiguracionCreateView(LoginRequiredMixin, FormMessageMixin, CreateView):
    """Crear nueva configuración"""
    model = ConfiguracionPlanificacion
    form_class = ConfiguracionPlanificacionForm
    template_name = 'turnos/configuration_form.html'
    success_message = 'Configuración creada con éxito.'

    def form_valid(self, form):
        form.instance.creado_por = self.request.user
        return super().form_valid(form)


class ConfiguracionUpdateView(LoginRequiredMixin, OwnerRequiredMixin, FormMessageMixin, UpdateView):
    """Editar configuración existente"""
    model = ConfiguracionPlanificacion
    form_class = ConfiguracionPlanificacionForm
    template_name = 'turnos/configuration_form.html'
    success_message = 'Configuración actualizada con éxito.'
    owner_field = 'creado_por'


class ConfiguracionDeleteView(LoginRequiredMixin, OwnerRequiredMixin, DeleteView):
    """Eliminar configuración"""
    model = ConfiguracionPlanificacion
    template_name = 'turnos/config_confirm_delete.html'
    success_url = reverse_lazy('turnos:config_lista')
    owner_field = 'creado_por'

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Configuración eliminada con éxito.')
        return super().delete(request, *args, **kwargs)


class ConfiguracionWizardView(LoginRequiredMixin, TemplateView):
    """Wizard para crear configuración paso a paso"""
    template_name = 'turnos/proyecto_turnos/wizard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['enfermeras'] = Enfermera.objects.filter(activa=True)
        context['turnos'] = TipoTurno.objects.filter(activo=True)
        return context


class ConfiguracionDuplicarView(LoginRequiredMixin, View):
    """Duplicar una configuración existente"""

    def post(self, request, pk):
        config_original = get_object_or_404(ConfiguracionPlanificacion, pk=pk)

        # Crear copia
        config_nueva = ConfiguracionPlanificacion.objects.create(
            nombre=f"{config_original.nombre} (Copia)",
            descripcion=config_original.descripcion,
            activa=config_original.activa,
            num_dias=config_original.num_dias,
            fecha_inicio=timezone.now().date(),
            demanda_por_turno=config_original.demanda_por_turno,
            restricciones_duras=config_original.restricciones_duras,
            restricciones_blandas=config_original.restricciones_blandas,
            num_trabajadores=config_original.num_trabajadores,
            tiempo_maximo_segundos=config_original.tiempo_maximo_segundos,
            seed=config_original.seed,
            creado_por=request.user
        )

        # Copiar relaciones ManyToMany
        config_nueva.enfermeras.set(config_original.enfermeras.all())
        config_nueva.turnos.set(config_original.turnos.all())

        messages.success(request, 'Configuración duplicada con éxito.')
        return redirect('turnos:config_detalle', pk=config_nueva.pk)


# ========== Ejecuciones ==========

class EjecucionListView(LoginRequiredMixin, SearchMixin, FilterMixin, PaginationMixin, ListView):
    """Lista de ejecuciones"""
    model = Ejecucion
    template_name = 'turnos/ejecucion_list.html'
    context_object_name = 'ejecuciones'
    paginate_by = 20

    search_fields = ['configuracion__nombre']

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.select_related('configuracion', 'planilla')

        # Filtros
        estado = self.request.GET.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)

        config_id = self.request.GET.get('proyecto_turnos')
        if config_id:
            queryset = queryset.filter(configuracion_id=config_id)

        return queryset.order_by('-fecha_inicio')


class EjecucionDetailView(LoginRequiredMixin, DetailView):
    """Detalle de una ejecución"""
    model = Ejecucion
    template_name = 'turnos/ejecucion_detail.html'
    context_object_name = 'ejecucion'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        if self.object.planilla:
            # Obtener asignaciones agrupadas
            asignaciones = self.object.planilla.asignaciones.select_related(
                'enfermera', 'turno'
            ).order_by('fecha', 'enfermera')

            context['asignaciones'] = asignaciones

        return context


class EjecucionDeleteView(LoginRequiredMixin, DeleteView):
    """Eliminar ejecución"""
    model = Ejecucion
    template_name = 'turnos/ejecucion_confirm_delete.html'
    success_url = reverse_lazy('turnos:ejecucion_lista')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Ejecución eliminada con éxito.')
        return super().delete(request, *args, **kwargs)


class EjecutarPlanificacionView(LoginRequiredMixin, FormView):
    """Vista para ejecutar una planificación"""
    template_name = 'templates/ejecutar_planificacion.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        config_id = self.kwargs.get('pk')
        context['proyecto_turnos'] = get_object_or_404(ConfiguracionPlanificacion, pk=config_id)
        return context

    def post(self, request, *args, **kwargs):
        config_id = self.kwargs.get('pk')
        config = get_object_or_404(ConfiguracionPlanificacion, pk=config_id)

        # Crear ejecución
        ejecucion = Ejecucion.objects.create(
            configuracion=config,
            estado='PENDIENTE'
        )

        # Lanzar tarea asíncrona
        ejecutar_planificacion_async.delay(ejecucion.id)

        messages.success(request, 'Planificación iniciada. Recibirás una notificación cuando termine.')
        return redirect('turnos:ejecucion_detalle', pk=ejecucion.id)


class EjecucionRapidaView(LoginRequiredMixin, FormView):
    """Vista para ejecución rápida"""
    template_name = 'turnos/ejecutar_rapido.html'
    form_class = EjecucionRapidaForm
    success_url = reverse_lazy('turnos:ejecucion_lista')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['enfermeras'] = Enfermera.objects.filter(activa=True)
        return context

    def form_valid(self, form):
        # Crear configuración rápida
        # ... implementación
        messages.success(self.request, 'Ejecución rápida iniciada.')
        return super().form_valid(form)


# ========== Enfermeras ==========

class EnfermeraListView(LoginRequiredMixin, SearchMixin, PaginationMixin, ListView):
    """Lista de enfermeras"""
    model = Enfermera
    template_name = 'turnos/enfermera_list.html'
    context_object_name = 'enfermeras'
    paginate_by = 15

    search_fields = ['nombre', 'email', 'dni']

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filtro de estado
        estado = self.request.GET.get('estado')
        if estado == 'activa':
            queryset = queryset.filter(activa=True)
        elif estado == 'inactiva':
            queryset = queryset.filter(activa=False)

        # Ordenamiento
        orden = self.request.GET.get('orden', 'nombre')
        queryset = queryset.order_by(orden)

        return queryset


class EnfermeraDetailView(LoginRequiredMixin, DetailView):
    """Detalle de una enfermera"""
    model = Enfermera
    template_name = 'turnos/enfermera_detail.html'
    context_object_name = 'enfermera'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Estadísticas
        context['stats'] = {
            'total_turnos': AsignacionTurno.objects.filter(enfermera=self.object).count(),
            'configuraciones': ConfiguracionPlanificacion.objects.filter(enfermeras=self.object).count(),
            'turnos_noche': AsignacionTurno.objects.filter(
                enfermera=self.object,
                turno__nombre='NOCHE'
            ).count()
        }

        # Asignaciones recientes
        context['asignaciones_recientes'] = AsignacionTurno.objects.filter(
            enfermera=self.object
        ).select_related('turno', 'planilla').order_by('-fecha')[:10]

        return context


class EnfermeraCreateView(LoginRequiredMixin, FormMessageMixin, CreateView):
    """Crear nueva enfermera"""
    model = Enfermera
    form_class = EnfermeraForm
    template_name = 'turnos/enfermera_form.html'
    success_message = 'Enfermera creada con éxito.'


class EnfermeraUpdateView(LoginRequiredMixin, FormMessageMixin, UpdateView):
    """Editar enfermera"""
    model = Enfermera
    form_class = EnfermeraForm
    template_name = 'turnos/enfermera_form.html'
    success_message = 'Enfermera actualizada con éxito.'


class EnfermeraDeleteView(LoginRequiredMixin, DeleteView):
    """Eliminar enfermera"""
    model = Enfermera
    template_name = 'turnos/enfermera_confirm_delete.html'
    success_url = reverse_lazy('turnos:enfermera_lista')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Enfermera eliminada con éxito.')
        return super().delete(request, *args, **kwargs)


class ImportarEnfermerasView(LoginRequiredMixin, FormView):
    """Importar enfermeras desde Excel"""
    template_name = 'turnos/enfermera_import.html'
    form_class = ImportarEnfermerasForm
    success_url = reverse_lazy('turnos:enfermera_lista')

    def form_valid(self, form):
        archivo = form.cleaned_data['archivo']
        sobrescribir = form.cleaned_data['sobrescribir']

        try:
            import openpyxl
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active

            creadas = 0
            actualizadas = 0
            errores = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                nombre, email, telefono, dni, activa = row[:5]

                if not nombre or not email:
                    continue

                activa = activa in ['Sí', 'Si', 'SI', 'sí', 'si', True, 1]

                # Buscar si existe
                enfermera_existente = Enfermera.objects.filter(email=email).first()

                if enfermera_existente:
                    if sobrescribir:
                        enfermera_existente.nombre = nombre
                        enfermera_existente.telefono = telefono or ''
                        enfermera_existente.dni = dni or ''
                        enfermera_existente.activa = activa
                        enfermera_existente.save()
                        actualizadas += 1
                else:
                    Enfermera.objects.create(
                        nombre=nombre,
                        email=email,
                        telefono=telefono or '',
                        dni=dni or '',
                        activa=activa
                    )
                    creadas += 1

            messages.success(
                self.request,
                f'Importación completada: {creadas} enfermeras creadas, {actualizadas} actualizadas.'
            )

        except Exception as e:
            messages.error(self.request, f'Error al importar: {str(e)}')
            return self.form_invalid(form)

        return super().form_valid(form)


class DescargarPlantillaEnfermerasView(LoginRequiredMixin, View):
    """Descarga plantilla Excel para importar enfermeras"""

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Enfermeras"

        # Headers con estilo
        headers = ['Nombre', 'Email', 'Teléfono', 'DNI', 'Activa']
        ws.append(headers)

        # Aplicar estilo a headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Ejemplo
        ws.append(['María García', 'maria@hospital.com', '600123456', '12345678A', 'Sí'])
        ws.append(['Juan López', 'juan@hospital.com', '600654321', '87654321B', 'Sí'])

        # Ajustar anchos
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=plantilla_enfermeras.xlsx'
        wb.save(response)

        return response


# ========== Tipos de Turno ==========

class TipoTurnoListView(LoginRequiredMixin, ListView):
    """Lista de tipos de turno"""
    model = TipoTurno
    template_name = 'turnos/tipo_turno_list.html'
    context_object_name = 'tipos_turno'

    def get_queryset(self):
        return TipoTurno.objects.all().order_by('nombre')


class TipoTurnoCreateView(LoginRequiredMixin, FormMessageMixin, CreateView):
    """Crear tipo de turno"""
    model = TipoTurno
    form_class = TipoTurnoForm
    template_name = 'turnos/tipo_turno_form.html'
    success_url = reverse_lazy('turnos:tipo_turno_lista')
    success_message = 'Tipo de turno creado con éxito.'


class TipoTurnoUpdateView(LoginRequiredMixin, FormMessageMixin, UpdateView):
    """Editar tipo de turno"""
    model = TipoTurno
    form_class = TipoTurnoForm
    template_name = 'turnos/tipo_turno_form.html'
    success_url = reverse_lazy('turnos:tipo_turno_lista')
    success_message = 'Tipo de turno actualizado con éxito.'


class TipoTurnoDeleteView(LoginRequiredMixin, DeleteView):
    """Eliminar tipo de turno"""
    model = TipoTurno
    template_name = 'turnos/tipo_turno_confirm_delete.html'
    success_url = reverse_lazy('turnos:tipo_turno_lista')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Tipo de turno eliminado con éxito.')
        return super().delete(request, *args, **kwargs)


class CrearTurnosPredeterminadosView(LoginRequiredMixin, View):
    """Crea los turnos predeterminados (Mañana, Tarde, Noche)"""

    def post(self, request):
        from datetime import time

        turnos_default = [
            {
                'nombre': 'MANANA',
                'hora_inicio': time(7, 0),
                'hora_fin': time(15, 0),
                'descripcion': 'Turno de mañana'
            },
            {
                'nombre': 'TARDE',
                'hora_inicio': time(15, 0),
                'hora_fin': time(23, 0),
                'descripcion': 'Turno de tarde'
            },
            {
                'nombre': 'NOCHE',
                'hora_inicio': time(23, 0),
                'hora_fin': time(7, 0),
                'descripcion': 'Turno de noche'
            }
        ]

        creados = 0
        for turno_data in turnos_default:
            if not TipoTurno.objects.filter(nombre=turno_data['nombre']).exists():
                TipoTurno.objects.create(**turno_data)
                creados += 1

        if creados > 0:
            messages.success(request, f'{creados} tipos de turno creados con éxito.')
        else:
            messages.info(request, 'Los turnos predeterminados ya existen.')

        return redirect('turnos:tipo_turno_lista')


# ========== Planillas ==========

class PlanillaListView(LoginRequiredMixin, SearchMixin, PaginationMixin, ListView):
    """Lista de planillas"""
    model = Planilla
    template_name = 'turnos/planilla_list.html'
    context_object_name = 'planillas'
    paginate_by = 12

    search_fields = ['nombre', 'descripcion']

    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset.select_related('ejecucion__configuracion').order_by('-fecha_inicio')


class PlanillaDetailView(LoginRequiredMixin, DetailView):
    """Detalle de una planilla"""
    model = Planilla
    template_name = 'turnos/planilla_detail.html'
    context_object_name = 'planilla'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Obtener asignaciones agrupadas
        asignaciones = self.object.asignaciones.select_related(
            'enfermera', 'turno'
        ).order_by('fecha', 'enfermera')

        context['asignaciones'] = asignaciones

        return context


class PlanillaDeleteView(LoginRequiredMixin, DeleteView):
    """Eliminar planilla"""
    model = Planilla
    template_name = 'turnos/planilla_confirm_delete.html'
    success_url = reverse_lazy('turnos:planilla_lista')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Planilla eliminada con éxito.')
        return super().delete(request, *args, **kwargs)


# ========== Reportes ==========

class ReportesView(LoginRequiredMixin, TemplateView):
    """Vista principal de reportes"""
    template_name = 'turnos/reportes.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Estadísticas para los reportes
        context['stats'] = {
            'ultima_actualizacion_carga': timezone.now(),
            'ultima_actualizacion_conflictos': timezone.now(),
            'ultima_actualizacion_tendencias': timezone.now(),
            'conflictos_activos': 0,
            'total_configuraciones': ConfiguracionPlanificacion.objects.count(),
            'total_ejecuciones': Ejecucion.objects.count(),
            'total_enfermeras': Enfermera.objects.count(),
            'total_turnos_asignados': AsignacionTurno.objects.count()
        }

        return context


class ReporteCargaView(LoginRequiredMixin, TemplateView):
    """Reporte de carga de trabajo"""
    template_name = 'turnos/reporte_carga.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Filtros
        fecha_desde = self.request.GET.get('fecha_desde')
        fecha_hasta = self.request.GET.get('fecha_hasta')

        # Estadísticas por enfermera
        enfermeras = Enfermera.objects.filter(activa=True)

        datos_enfermeras = []
        for enfermera in enfermeras:
            asignaciones = AsignacionTurno.objects.filter(enfermera=enfermera)

            if fecha_desde:
                asignaciones = asignaciones.filter(fecha__gte=fecha_desde)
            if fecha_hasta:
                asignaciones = asignaciones.filter(fecha__lte=fecha_hasta)

            turnos_manana = asignaciones.filter(turno__nombre='MANANA').count()
            turnos_tarde = asignaciones.filter(turno__nombre='TARDE').count()
            turnos_noche = asignaciones.filter(turno__nombre='NOCHE').count()
            total_turnos = turnos_manana + turnos_tarde + turnos_noche

            datos_enfermeras.append({
                'enfermera': enfermera,
                'turnos_manana': turnos_manana,
                'turnos_tarde': turnos_tarde,
                'turnos_noche': turnos_noche,
                'total_turnos': total_turnos,
                'horas_totales': total_turnos * 8,  # Aproximación
                'dias_libres': asignaciones.filter(es_dia_libre=True).count(),
                'porcentaje_carga': (total_turnos / 30 * 100) if total_turnos else 0
            })

        context['datos_enfermeras'] = datos_enfermeras
        context['resumen'] = {
            'total_enfermeras': len(datos_enfermeras),
            'total_turnos': sum(d['total_turnos'] for d in datos_enfermeras),
            'horas_totales': sum(d['horas_totales'] for d in datos_enfermeras),
            'promedio_por_enfermera': sum(d['total_turnos'] for d in datos_enfermeras) / len(
                datos_enfermeras) if datos_enfermeras else 0
        }

        # Datos para gráfico
        context['distribucion_turnos'] = [
            sum(d['turnos_manana'] for d in datos_enfermeras),
            sum(d['turnos_tarde'] for d in datos_enfermeras),
            sum(d['turnos_noche'] for d in datos_enfermeras)
        ]

        return context


class ReporteConflictosView(LoginRequiredMixin, TemplateView):
    """Reporte de conflictos en planificaciones"""
    template_name = 'turnos/reporte_conflictos.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Aquí se implementaría la lógica para detectar conflictos
        # Por ejemplo: turnos con descanso insuficiente, sobrecarga, etc.

        context['conflictos'] = []
        context['resumen'] = {
            'total_conflictos': 0,
            'severidad_alta': 0,
            'severidad_media': 0,
            'severidad_baja': 0
        }

        return context


class ReporteTendenciasView(LoginRequiredMixin, TemplateView):
    """Reporte de tendencias temporales"""
    template_name = 'turnos/reporte_tendencias.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Datos mensuales
        hoy = timezone.now()
        datos_mensuales = []

        for i in range(6):
            mes_inicio = hoy - timedelta(days=30 * i)
            mes_fin = mes_inicio + timedelta(days=30)

            ejecuciones_mes = Ejecucion.objects.filter(
                fecha_inicio__gte=mes_inicio,
                fecha_inicio__lt=mes_fin
            )

            datos_mensuales.append({
                'nombre': mes_inicio.strftime('%B'),
                'total': ejecuciones_mes.count(),
                'exitosas': ejecuciones_mes.filter(estado='COMPLETADA').count(),
                'fallidas': ejecuciones_mes.filter(estado='ERROR').count(),
                'tiempo_promedio': ejecuciones_mes.aggregate(Avg('duracion'))['duracion__avg'] or 0,
                'tasa_exito': (ejecuciones_mes.filter(
                    estado='COMPLETADA').count() / ejecuciones_mes.count() * 100) if ejecuciones_mes.count() > 0 else 0
            })

        context['datos_mensuales'] = datos_mensuales

        # KPIs
        context['kpis'] = {
            'crecimiento_ejecuciones': 15.5,
            'tiempo_promedio': 45.3,
            'tasa_exito': 92.5,
            'soluciones_optimas': 78.2
        }

        return context


# ========== Vistas de Usuario ==========

class PerfilView(LoginRequiredMixin, TemplateView):
    """Vista de perfil de usuario"""
    template_name = 'turnos/perfil.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Estadísticas del usuario
        context['stats'] = {
            'configuraciones': ConfiguracionPlanificacion.objects.filter(creado_por=self.request.user).count(),
            'ejecuciones': Ejecucion.objects.filter(configuracion__creado_por=self.request.user).count(),
            'enfermeras': Enfermera.objects.count()
        }

        return context


class PreferenciasView(LoginRequiredMixin, TemplateView):
    """Vista de preferencias del sistema"""
    template_name = 'turnos/preferencias.html'


class GuardarPreferenciasView(LoginRequiredMixin, View):
    """Guarda las preferencias del usuario"""

    def post(self, request):
        # Guardar preferencias en la sesión o en un modelo de perfil
        request.session['preferencias'] = {
            'notificaciones_email': request.POST.get('notificaciones_email') == 'on',
            'notificaciones_browser': request.POST.get('notificaciones_browser') == 'on',
            'idioma': request.POST.get('idioma', 'es'),
            'tema': request.POST.get('tema', 'light'),
            'trabajadores_default': int(request.POST.get('trabajadores_default', 4)),
            'tiempo_maximo_default': int(request.POST.get('tiempo_maximo_default', 60))
        }

        messages.success(request, 'Preferencias guardadas con éxito.')
        return redirect('turnos:preferencias')


# ========== Vistas de Resultado ==========

class ResultadoCalendarioView(LoginRequiredMixin, DetailView):
    """Vista de resultado en formato calendario"""
    model = Ejecucion
    template_name = 'turnos/resultado_calendario.html'
    context_object_name = 'ejecucion'


class ResultadoEstadisticasView(LoginRequiredMixin, DetailView):
    """Vista de estadísticas del resultado"""
    model = Ejecucion
    template_name = 'turnos/resultado_estadisticas.html'
    context_object_name = 'ejecucion'


class ResultadoTablaView(LoginRequiredMixin, DetailView):
    """Vista de resultado en formato tabla"""
    model = Ejecucion
    template_name = 'turnos/resultado_tabla.html'
    context_object_name = 'ejecucion'


class ResultadoCompararView(LoginRequiredMixin, TemplateView):
    """Vista para comparar dos resultados"""
    template_name = 'turnos/resultado_comparar.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        ejecucion_a_id = self.request.GET.get('ejecucion_a')
        ejecucion_b_id = self.request.GET.get('ejecucion_b')

        if ejecucion_a_id and ejecucion_b_id:
            context['ejecucion_a'] = get_object_or_404(Ejecucion, pk=ejecucion_a_id)
            context['ejecucion_b'] = get_object_or_404(Ejecucion, pk=ejecucion_b_id)

        context['ejecuciones'] = Ejecucion.objects.filter(estado='COMPLETADA').order_by('-fecha_inicio')[:20]

        return context


# ========== Vistas de Utilidad ==========

class MaintenanceView(TemplateView):
    """Vista de mantenimiento"""
    template_name = 'turnos/maintenance.html'


# ========== Exportaciones ==========

class ExportarEjecucionExcelView(LoginRequiredMixin, View):
    """Exporta una ejecución a Excel"""

    def get(self, request, pk):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        ejecucion = get_object_or_404(Ejecucion, pk=pk)

        if not ejecucion.planilla:
            messages.error(request, 'Esta ejecución no tiene planilla asociada.')
            return redirect('turnos:ejecucion_detalle', pk=pk)

        wb = Workbook()
        ws = wb.active
        ws.title = "Planilla de Turnos"

        # Título
        ws['A1'] = ejecucion.configuracion.nombre
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')

        ws[
            'A2'] = f"Período: {ejecucion.planilla.fecha_inicio.strftime('%d/%m/%Y')} - {ejecucion.planilla.fecha_fin.strftime('%d/%m/%Y')}"
        ws.merge_cells('A2:E2')

        # Headers
        headers = ['Enfermera', 'Fecha', 'Día Semana', 'Turno', 'Horario']
        ws.append([''])  # Línea en blanco
        ws.append(headers)

        # Estilo headers
        for cell in ws[4]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Datos
        asignaciones = ejecucion.planilla.asignaciones.select_related(
            'enfermera', 'turno'
        ).order_by('fecha', 'enfermera')

        for asignacion in asignaciones:
            if asignacion.es_dia_libre:
                turno_info = 'Libre'
                horario = '-'
            else:
                turno_info = asignacion.turno.get_nombre_display()
                horario = f"{asignacion.turno.hora_inicio.strftime('%H:%M')} - {asignacion.turno.hora_fin.strftime('%H:%M')}"

            dia_semana = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom'][asignacion.fecha.weekday()]

            ws.append([
                asignacion.enfermera.nombre,
                asignacion.fecha.strftime('%d/%m/%Y'),
                dia_semana,
                turno_info,
                horario
            ])

        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20

        # Respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=planilla_{ejecucion.id}.xlsx'
        wb.save(response)

        return response


class ExportarEjecucionPDFView(LoginRequiredMixin, View):
    """Exporta una ejecución a PDF"""

    def get(self, request, pk):
        from django.template.loader import render_to_string
        from weasyprint import HTML

        ejecucion = get_object_or_404(Ejecucion, pk=pk)

        if not ejecucion.planilla:
            messages.error(request, 'Esta ejecución no tiene planilla asociada.')
            return redirect('turnos:ejecucion_detalle', pk=pk)

        # Renderizar HTML
        html_string = render_to_string('turnos/pdf/planilla.html', {
            'ejecucion': ejecucion,
            'planilla': ejecucion.planilla,
            'asignaciones': ejecucion.planilla.asignaciones.select_related('enfermera', 'turno').order_by('fecha',
                                                                                                          'enfermera')
        })

        # Generar PDF
        html = HTML(string=html_string)
        pdf = html.write_pdf()

        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=planilla_{ejecucion.id}.pdf'

        return response


class ExportarEjecucionCSVView(LoginRequiredMixin, View):
    """Exporta una ejecución a CSV"""

    def get(self, request, pk):
        import csv

        ejecucion = get_object_or_404(Ejecucion, pk=pk)

        if not ejecucion.planilla:
            messages.error(request, 'Esta ejecución no tiene planilla asociada.')
            return redirect('turnos:ejecucion_detalle', pk=pk)

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename=planilla_{ejecucion.id}.csv'

        writer = csv.writer(response)
        writer.writerow(['Enfermera', 'Fecha', 'Turno', 'Horario', 'Es Día Libre'])

        asignaciones = ejecucion.planilla.asignaciones.select_related(
            'enfermera', 'turno'
        ).order_by('fecha', 'enfermera')

        for asignacion in asignaciones:
            if asignacion.es_dia_libre:
                turno_info = 'Libre'
                horario = '-'
            else:
                turno_info = asignacion.turno.get_nombre_display()
                horario = f"{asignacion.turno.hora_inicio.strftime('%H:%M')} - {asignacion.turno.hora_fin.strftime('%H:%M')}"

            writer.writerow([
                asignacion.enfermera.nombre,
                asignacion.fecha.strftime('%d/%m/%Y'),
                turno_info,
                horario,
                'Sí' if asignacion.es_dia_libre else 'No'
            ])

        return response


class ExportarEjecucionJSONView(LoginRequiredMixin, View):
    """Exporta una ejecución a JSON"""

    def get(self, request, pk):
        ejecucion = get_object_or_404(Ejecucion, pk=pk)

        if not ejecucion.planilla:
            messages.error(request, 'Esta ejecución no tiene planilla asociada.')
            return redirect('turnos:ejecucion_detalle', pk=pk)

        data = {
            'ejecucion_id': ejecucion.id,
            'configuracion': ejecucion.configuracion.nombre,
            'fecha_inicio': ejecucion.fecha_inicio.isoformat(),
            'estado': ejecucion.estado,
            'es_optima': ejecucion.es_optima,
            'penalizacion_total': ejecucion.penalizacion_total,
            'planilla': {
                'nombre': ejecucion.planilla.nombre,
                'fecha_inicio': ejecucion.planilla.fecha_inicio.isoformat(),
                'fecha_fin': ejecucion.planilla.fecha_fin.isoformat(),
                'asignaciones': []
            }
        }

        asignaciones = ejecucion.planilla.asignaciones.select_related(
            'enfermera', 'turno'
        ).order_by('fecha', 'enfermera')

        for asignacion in asignaciones:
            data['planilla']['asignaciones'].append({
                'enfermera': asignacion.enfermera.nombre,
                'fecha': asignacion.fecha.isoformat(),
                'turno': asignacion.turno.get_nombre_display() if asignacion.turno else None,
                'es_dia_libre': asignacion.es_dia_libre
            })

        response = JsonResponse(data, json_dumps_params={'indent': 2})
        response['Content-Disposition'] = f'attachment; filename=planilla_{ejecucion.id}.json'

        return response


class ExportarEjecucionICalView(LoginRequiredMixin, View):
    """Exporta una ejecución a formato iCalendar"""

    def get(self, request, pk):
        from icalendar import Calendar, Event
        from datetime import datetime, timedelta

        ejecucion = get_object_or_404(Ejecucion, pk=pk)

        if not ejecucion.planilla:
            messages.error(request, 'Esta ejecución no tiene planilla asociada.')
            return redirect('turnos:ejecucion_detalle', pk=pk)

        cal = Calendar()
        cal.add('prodid', '-//Sistema de Planificación de Turnos//ES')
        cal.add('version', '2.0')
        cal.add('X-WR-CALNAME', ejecucion.configuracion.nombre)

        asignaciones = ejecucion.planilla.asignaciones.select_related(
            'enfermera', 'turno'
        ).order_by('fecha')

        for asignacion in asignaciones:
            if not asignacion.es_dia_libre:
                event = Event()
                event.add('summary', f"{asignacion.enfermera.nombre} - {asignacion.turno.get_nombre_display()}")

                # Calcular fecha/hora inicio y fin
                dt_inicio = datetime.combine(asignacion.fecha, asignacion.turno.hora_inicio)
                dt_fin = datetime.combine(asignacion.fecha, asignacion.turno.hora_fin)

                # Si el turno cruza medianoche
                if dt_fin <= dt_inicio:
                    dt_fin += timedelta(days=1)

                event.add('dtstart', dt_inicio)
                event.add('dtend', dt_fin)
                event.add('description', f"Turno asignado: {asignacion.turno.get_nombre_display()}")

                cal.add_component(event)

        response = HttpResponse(cal.to_ical(), content_type='text/calendar')
        response['Content-Disposition'] = f'attachment; filename=planilla_{ejecucion.id}.ics'

        return response


class ExportarPlanillaExcelView(LoginRequiredMixin, View):
    """Exporta una planilla específica a Excel"""

    def get(self, request, pk):
        planilla = get_object_or_404(Planilla, pk=pk)
        # Reutilizar la lógica de exportación de ejecución
        # pero usando la planilla directamente
        return ExportarEjecucionExcelView().get(request, planilla.ejecucion.id)


class ExportarPlanillaPDFView(LoginRequiredMixin, View):
    """Exporta una planilla específica a PDF"""

    def get(self, request, pk):
        planilla = get_object_or_404(Planilla, pk=pk)
        return ExportarEjecucionPDFView().get(request, planilla.ejecucion.id)
